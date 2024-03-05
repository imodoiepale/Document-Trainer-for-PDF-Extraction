import tkinter as tk
from tkinter import filedialog, OptionMenu, ttk
import pdfplumber
import json
import openpyxl
import os
from datetime import datetime
import subprocess
import re

def load_positions(filename):
    # Load the positions from the selected file
    with open(filename, 'r') as f:
        positions = json.load(f)
    return positions

def extract_text():
    positions = load_positions(root.filename.get())

    filenames = filedialog.askopenfilenames(
        title="Select PDF Files",
        filetypes=[("PDF Files", "*.pdf")]
    )

    if filenames:
        total_files = len(filenames)
        progress['maximum'] = total_files
        progress['value'] = 0
        progress.pack(side=tk.BOTTOM, fill=tk.X)
        root.update_idletasks()

        # Get current date and time
        current_datetime = datetime.now().strftime("%H.%M-%d.%m.%Y")

        # Create a new workbook and select the active sheet
        wb = openpyxl.Workbook()
        sheet = wb.active

        # Write the positions extraction headers
        positions_headers = list(positions.keys())
        sheet.append(positions_headers)

        # Write the table extraction headers starting from column J
        table_headers = ['NAME', 'DESCRIPTION', 'ADDRESS', 'NATIONALITY', 'SHARES']
        for col_index, header in enumerate(table_headers, start=10):
            sheet.cell(row=1, column=col_index, value=header)

        # Initialize row index for table data
        table_row_index = 2

        for index, file in enumerate(filenames, start=1):
            with pdfplumber.open(file) as pdf:
                for page in pdf.pages:
                    # Extract positions data
                    position_row = []
                    for label, position in positions.items():
                        # Crop the page to the position
                        cropped_page = page.crop((position['x1'], position['y1'], position['x2'], position['y2']))
                        # Extract the text from the cropped page
                        text = cropped_page.extract_text()
                        # Remove commas from the text
                        text = text.replace(",", "")
                        # Check if the text is numerical
                        if re.match(r'^-?\d+(?:\.\d+)?$', text):
                            text = float(text)
                        # Add the extracted text to the row
                        position_row.append(text)
                    # Append positions data to the sheet
                    sheet.append(position_row)

                    # Extract tables
                    table = page.extract_table()
                    if table:
                        # Write the table data starting from column J
                        for table_row in table[1:]:
                            if table_row[0] != 'TOTAL':  # Exclude the "TOTAL" row
                                table_row_index = sheet.max_row + 1  # Increment the row index
                                for col_index, value in enumerate(table_row, start=10):
                                    sheet.cell(row=table_row_index, column=col_index, value=value)
                                # Repeat company name in the same row under the company column
                                company_name = position_row[0]
                                if company_name:
                                    sheet.cell(row=table_row_index, column=1, value=company_name)
                                # Increment the row index for the next table row
                                table_row_index += 1
                                

                        # Skip two rows
                        sheet.append([''] * sheet.max_column)  # Empty row
                        sheet.append([''] * sheet.max_column)  # Empty row

            progress['value'] = index
            root.update_idletasks()

        # Save the workbook to an Excel file with the current date and time
        file_path = f"CR12 - EXTRACTION _{current_datetime}.xlsx"
        wb.save(file_path)
        progress.pack_forget()
        complete_label.pack()
        file_label['text'] = "Open Excel file: " + os.path.abspath(file_path)
        file_label['command'] = lambda: open_file_explorer(file_path)
        file_label.pack()

def open_file_explorer(file_path):
    subprocess.Popen('explorer /select,"{}"'.format(os.path.abspath(file_path)))

root = tk.Tk()
root.title("PDF Text Extractor")
root.geometry("600x400")

# Get a list of all text files in the current directory
txt_files = [f for f in os.listdir() if f.endswith('.txt')]

if txt_files:
    # Create a StringVar to hold the selected filename
    root.filename = tk.StringVar(root)
    root.filename.set(txt_files[0])  # Set the default value to the first file

    # Create the dropdown menu
    dropdown = OptionMenu(root, root.filename, *txt_files)
    dropdown.pack()

    extract_button = tk.Button(root, text="Select Files and Extract Text", command=extract_text)
    extract_button.pack()
else:
    # Display a message if no text files are found
    label = tk.Label(root, text="No TXT files found in the directory.")
    label.pack()

progress = ttk.Progressbar(root, orient=tk.HORIZONTAL, length=200, mode='determinate')
complete_label = tk.Label(root, text="Extraction complete", fg="green")
file_label = tk.Button(root, text="Open Excel file: ", fg="blue", cursor="hand2", command=lambda: open_file_explorer(None))

root.mainloop()
